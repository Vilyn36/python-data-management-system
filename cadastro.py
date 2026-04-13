"""
Módulo de cadastro para a Clínica VIDA+

Fluxo de persistência (atual):
- Ao iniciar, o módulo tenta carregar os dados de uma planilha mestre XLSX em
    `~/Documentos/LISTA + VIDA.xlsx` (constante `MASTER_XLSX`). Para isso usa
    `openpyxl` quando disponível.
- Se a planilha mestre não existir ou `openpyxl` não estiver disponível, o
    sistema carrega os dados a partir de arquivos CSV locais em `data/`.
- Sempre que houver um novo cadastro (paciente, médico, exame) ou um
    agendamento, os CSVs são atualizados e a planilha mestre XLSX é gravada
    automaticamente (função `save_master_xlsx()`). Se `openpyxl` não estiver
    instalado, os CSVs continuam sendo atualizados e o usuário será avisado.

Notas:
- As abas esperadas na planilha mestre são: 'pacientes', 'medicos', 'exames'
    e 'agendamentos' com cabeçalhos específicos. Em `pacientes` agora são
    esperadas as colunas ['id', 'nome', 'idade', 'telefone'], onde `id` é um
    identificador numérico de 5 dígitos (zero à esquerda). O carregamento
    continuará aceitando arquivos antigos (sem a coluna de ID), mas validará os
    cabeçalhos básicos antes de aceitar o conteúdo.
"""

from datetime import datetime, timedelta
import unicodedata
import os
import csv
import shutil
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    HAVE_OPENPYXL = True
except Exception:
    HAVE_OPENPYXL = False

# pasta onde os CSVs serão salvos
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(DATA_DIR, exist_ok=True)

# caminho da planilha mestre (arquivo .xlsx no diretório Documentos do usuário)
MASTER_XLSX = os.path.join(os.path.expanduser(
    "~"), "Documentos", "LISTA + VIDA.xlsx")

pacientes = []
medicos = []
exames = []
agendamentos = []
disponibilidade = []
atendimentos = []
cadastros_atendimento = []
convenios = []

# nomes dos arquivos
PACIENTES_CSV = os.path.join(DATA_DIR, "pacientes.csv")
MEDICOS_CSV = os.path.join(DATA_DIR, "medicos.csv")
EXAMES_CSV = os.path.join(DATA_DIR, "exames.csv")
AGENDAMENTOS_CSV = os.path.join(DATA_DIR, "agendamentos.csv")
DISPO_CSV = os.path.join(DATA_DIR, "disponibilidade.csv")
ATENDIMENTOS_CSV = os.path.join(DATA_DIR, "atendimentos.csv")
CADASTROS_ATENDIMENTO_CSV = os.path.join(DATA_DIR, "cadastros_atendimento.csv")
CONVENIOS_CSV = os.path.join(DATA_DIR, "convenios.csv")

next_paciente_id = 1
DEFAULT_JANELA_PROXIMIDADE_MINUTOS = 120


def normalizar_texto(texto):
    if texto is None:
        return ""
    if not isinstance(texto, str):
        texto = str(texto)
    texto = unicodedata.normalize('NFKD', texto)
    texto = ''.join(ch for ch in texto if not unicodedata.combining(ch))
    return texto.strip().lower()


def validar_nome(nome):
    if not isinstance(nome, str):
        return False
    nome_limpo = nome.strip()
    if not nome_limpo:
        return False
    return all(ch.isalpha() or ch.isspace() for ch in nome_limpo)


def _converter_id_para_int(valor):
    try:
        texto = str(valor).strip()
    except Exception:
        return None
    if not texto:
        return None
    if texto.isdigit():
        return int(texto)
    return None


def _formatar_id_paciente(valor_int):
    if valor_int is None or valor_int < 1:
        return ""
    return f"{valor_int:05d}"


def sincronizar_ids_pacientes():
    global next_paciente_id
    maior_id = 0
    # normalizar IDs existentes e rastrear maior valor
    for paciente in pacientes:
        pid = _converter_id_para_int(paciente.get('id'))
        if pid is not None:
            paciente['id'] = _formatar_id_paciente(pid)
            if pid > maior_id:
                maior_id = pid
        else:
            paciente['id'] = ''

    next_paciente_id = maior_id + 1 if maior_id else 1

    for paciente in pacientes:
        if not paciente.get('id'):
            paciente['id'] = _formatar_id_paciente(next_paciente_id)
            next_paciente_id += 1


def gerar_id_paciente():
    global next_paciente_id
    if next_paciente_id < 1:
        next_paciente_id = 1
    novo_id_int = next_paciente_id
    next_paciente_id += 1
    return _formatar_id_paciente(novo_id_int)


def obter_id_legivel(paciente):
    pid = str(paciente.get('id', '') or '').strip()
    return pid if pid else "Sem ID"


def migrate_agendamentos():
    """Converte agendamentos com data_hora em string para datetime quando possível."""
    for a in agendamentos:
        dh = a.get('data_hora')
        if isinstance(dh, str):
            conv = validar_data_hora(dh)
            if conv:
                a['data_hora'] = conv


def migrate_atendimentos():
    """Converte históricos com data_hora em string para datetime quando possível."""
    for h in atendimentos:
        dh = h.get('data_hora')
        if isinstance(dh, str):
            conv = validar_data_hora(dh)
            if conv:
                h['data_hora'] = conv


def slots_disponiveis_medico(nome_medico, tipo):
    if not nome_medico:
        return []

    medico_norm = normalizar_texto(nome_medico)
    tipo_alvo = (tipo or '').strip().lower()
    candidatos = []
    for slot in disponibilidade:
        nome_slot = (slot.get('medico') or '').strip()
        if not nome_slot or normalizar_texto(nome_slot) != medico_norm:
            continue
        tipo_slot = (slot.get('tipo') or '').strip().lower()
        if tipo_alvo and tipo_slot and tipo_slot not in ('ambos', tipo_alvo):
            continue
        dh = slot.get('data_hora')
        if isinstance(dh, str):
            convertido = validar_data_hora(dh)
            if not convertido:
                continue
            dh = convertido
        if not isinstance(dh, datetime):
            continue
        candidatos.append({'data_hora': dh, 'tipo': tipo_slot, 'raw': slot})

    if not candidatos:
        return []

    ocupados = set()
    for ag in agendamentos:
        if normalizar_texto(ag.get('medico', '')) != medico_norm:
            continue
        existente = ag.get('data_hora')
        if isinstance(existente, str):
            existente = validar_data_hora(existente)
        if isinstance(existente, datetime):
            ocupados.add(existente)

    livres = [slot for slot in candidatos if slot['data_hora'] not in ocupados]
    livres.sort(key=lambda entry: entry['data_hora'])
    return livres


def selecionar_exame_interativo():
    if not exames:
        print("Erro: não há exames cadastrados. Cadastre exames antes de agendar.")
        return None

    while True:
        print("\n--- Exames Disponíveis ---")
        for idx, exame in enumerate(exames, start=1):
            tipo_exame = exame.get('tipo')
            tipo_info = f" | Tipo: {tipo_exame}" if tipo_exame else ""
            descricao = exame.get('descricao')
            descricao_info = f" | Descrição: {descricao}" if descricao else ""
            print(f"{idx}. {exame.get('nome', '')}{tipo_info}{descricao_info}")
        print("0. Voltar ao menu principal")
        escolha = input("Escolha o número do exame: ").strip()
        if escolha == '0':
            print("Voltando ao menu principal.\n")
            return None
        if escolha.isdigit():
            escolha_int = int(escolha)
            if 1 <= escolha_int <= len(exames):
                selecionado = exames[escolha_int - 1].get('nome', '')
                print(f"Exame selecionado: {selecionado}\n")
                return selecionado
        print("Seleção inválida. Tente novamente.")


def selecionar_paciente_interativo():
    if not pacientes:
        print("Erro: não há pacientes cadastrados.")
        return None

    sincronizar_ids_pacientes()

    while True:
        termo = input("Nome (ou parte do nome) do paciente: ").strip()
        if not termo:
            print("Informe ao menos uma letra para buscar o paciente.")
            continue

        termo_norm = normalizar_texto(termo)
        candidatos = [p for p in pacientes if termo_norm
                      in normalizar_texto(p.get('nome', '').strip())]
        if not candidatos:
            while True:
                print("1. Listar pacientes cadastrados")
                print("0. Voltar ao menu principal")
                acao = input("Escolha uma opção: ").strip()
                if acao == "0":
                    print("Voltando ao menu principal.\n")
                    return None
                if acao == "1":
                    if not pacientes:
                        print("Não há pacientes cadastrados.")
                        return None
                    print("\n--- Pacientes Cadastrados ---")
                    for idx, cand in enumerate(pacientes, start=1):
                        idade_info = cand.get('idade')
                        idade_str = f" | Idade: {idade_info}" if isinstance(
                            idade_info, (int, float)) else ""
                        pid = obter_id_legivel(cand)
                        print(
                            f"{idx}. [{pid}] {cand.get('nome', '')}{idade_str}")
                    print("V. Voltar")
                    print("0. Voltar ao menu principal")
                    escolha_lista = input(
                        "Escolha o número do paciente: ").strip().lower()
                    if escolha_lista == '0':
                        print("Voltando ao menu principal.\n")
                        return None
                    if escolha_lista == 'v':
                        print()
                        break
                    if escolha_lista.isdigit():
                        escolha_int = int(escolha_lista)
                        if 1 <= escolha_int <= len(pacientes):
                            print()
                            return pacientes[escolha_int - 1]
                    print("Seleção inválida. Tente novamente.")
                    continue
                print("Opção inválida. Tente novamente.")
            continue

        if len(candidatos) == 1:
            return candidatos[0]

        print("Pacientes encontrados:")
        for idx, cand in enumerate(candidatos, start=1):
            idade_info = cand.get('idade')
            idade_str = f" | Idade: {idade_info}" if isinstance(
                idade_info, (int, float)) else ""
            pid = obter_id_legivel(cand)
            print(f"{idx}. [{pid}] {cand.get('nome', '')}{idade_str}")
        print("V. Pesquisar novamente")
        print("0. Voltar ao menu principal")
        escolha = input("Escolha o número do paciente: ").strip().lower()
        if escolha == '0':
            print("Voltando ao menu principal.\n")
            return None
        if escolha == 'v':
            continue
        if escolha.isdigit():
            escolha_int = int(escolha)
            if 1 <= escolha_int <= len(candidatos):
                return candidatos[escolha_int - 1]
        print("Seleção inválida. Tente novamente.")


def localizar_paciente_por_identificadores(paciente_id=None, paciente_nome=None):
    pid = str(paciente_id or '').strip()
    nome_norm = normalizar_texto(paciente_nome) if paciente_nome else None
    for paciente in pacientes:
        atual_id = str(paciente.get('id', '') or '').strip()
        if pid and atual_id and atual_id == pid:
            return paciente
        if nome_norm and normalizar_texto(paciente.get('nome', '')) == nome_norm:
            return paciente
    return None


def avaliar_consulta_normal(tem_agendamento, documentos_ok, medico_disponivel, pagamentos_em_dia):
    return (tem_agendamento and documentos_ok and medico_disponivel) or \
        (documentos_ok and medico_disponivel and pagamentos_em_dia)


def avaliar_emergencia(documentos_ok, medico_disponivel, pagamentos_em_dia):
    return medico_disponivel and (documentos_ok or pagamentos_em_dia)


def obter_especialidade_por_medico(nome_medico):
    alvo_norm = normalizar_texto(nome_medico or "")
    for medico in medicos:
        if normalizar_texto(medico.get('nome', '')) == alvo_norm:
            return medico.get('especialidade') or "Não informado"
    return "Não informado"


def paciente_tem_agendamento_na_data(nome_paciente, data_alvo):
    if not nome_paciente or not data_alvo:
        return False
    if isinstance(data_alvo, datetime):
        data_referencia = data_alvo.date()
    else:
        data_referencia = data_alvo

    paciente_norm = normalizar_texto(nome_paciente)
    for agendamento in agendamentos:
        if normalizar_texto(agendamento.get('paciente', '')) != paciente_norm:
            continue
        dh = agendamento.get('data_hora')
        if isinstance(dh, str):
            dh = validar_data_hora(dh)
        if isinstance(dh, datetime) and dh.date() == data_referencia:
            return True
    return False


def filtrar_slots_por_proximidade(slots, referencia_dt, janela_td):
    if not referencia_dt:
        return slots
    filtrados = []
    for slot in slots:
        dh = slot.get('data_hora')
        if isinstance(dh, str):
            convertido = validar_data_hora(dh)
            if convertido:
                slot = dict(slot)
                slot['data_hora'] = convertido
                dh = convertido
        if not isinstance(dh, datetime):
            continue
        if janela_td is None:
            if dh.date() != referencia_dt.date():
                continue
            if dh >= referencia_dt:
                filtrados.append(slot)
        elif abs(dh - referencia_dt) <= janela_td:
            filtrados.append(slot)
    return filtrados


def listar_especialidades_disponiveis_para_data(data_referencia):
    if isinstance(data_referencia, datetime):
        ref_dt = data_referencia
    else:
        ref_dt = datetime.combine(data_referencia, datetime.min.time())

    print("\n--- Especialidades disponíveis ---")
    if not disponibilidade:
        print("Nenhum horário cadastrado para qualquer especialidade.")
        return

    especialidades = {}
    for slot in disponibilidade:
        dh = slot.get('data_hora')
        if isinstance(dh, str):
            dh = validar_data_hora(dh)
        if not isinstance(dh, datetime):
            continue
        if dh.date() != ref_dt.date():
            continue
        if dh < ref_dt:
            continue
        especialidade = obter_especialidade_por_medico(slot.get('medico', ''))
        especialidades.setdefault(especialidade, 0)
        especialidades[especialidade] += 1

    if not especialidades:
        print("Nenhuma especialidade possui horários a partir do horário do cadastro.")
        return

    print(
        f"Especialidades com horários a partir de {ref_dt.strftime('%d/%m/%Y %H:%M')}:")
    for esp, quantidade in sorted(especialidades.items()):
        print(f"- {esp} ({quantidade} horário(s))")
    print()


def format_telefone(telefone_raw: str) -> str:
    """Normaliza e formata o telefone antes de salvar.

    - Remove todos os caracteres não numéricos.
    - Se tiver 11 dígitos -> formata como (AA) AAAAA-AAAA
    - Se tiver 10 dígitos -> formata como (AA) AAAA-AAAA
    - Caso contrário retorna apenas os dígitos limpos.
    """
    if telefone_raw is None:
        return ""
    digits = ''.join(ch for ch in telefone_raw if ch.isdigit())

    # Formata de acordo com o número de dígitos sem adicionar DDD automaticamente:
    # 11 dígitos -> (AA) AAAAA-AAAA
    # 10 dígitos -> (AA) AAAA-AAAA
    # 9 dígitos -> ABBBB-CCCC (sem DDD) -> 5-4
    # 8 dígitos -> AAAA-BBBB (sem DDD) -> 4-4
    if len(digits) == 11:
        return f"({digits[0:2]}) {digits[2:7]}-{digits[7:11]}"
    if len(digits) == 10:
        return f"({digits[0:2]}) {digits[2:6]}-{digits[6:10]}"
    if len(digits) == 9:
        return f"{digits[0:5]}-{digits[5:9]}"
    if len(digits) == 8:
        return f"{digits[0:4]}-{digits[4:8]}"
    return digits


def formatar_cpf(cpf_raw: str) -> str:
    if cpf_raw is None:
        return ""
    digits = ''.join(ch for ch in cpf_raw if ch.isdigit())
    if len(digits) == 11:
        return f"{digits[0:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:11]}"
    return digits


def _write_csv(path, fieldnames, rows):
    with open(path, "w", newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)


def _read_csv(path):
    if not os.path.exists(path):
        return []
    with open(path, "r", newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        return [row for row in reader]


def save_pacientes():
    sincronizar_ids_pacientes()
    rows = []
    for p in pacientes:
        rows.append({
            "id": p.get("id", ""),
            "nome": p.get("nome", ""),
            "idade": str(p.get("idade", "")),
            "telefone": p.get("telefone", "")
        })
    _write_csv(PACIENTES_CSV, ["id", "nome", "idade", "telefone"], rows)


def save_medicos():
    rows = []
    for m in medicos:
        rows.append({
            "nome": m.get("nome", ""),
            "especialidade": m.get("especialidade", ""),
            "CRM": m.get("CRM", ""),
            "telefone": m.get("telefone", "")
        })
    _write_csv(MEDICOS_CSV, ["nome", "especialidade", "CRM", "telefone"], rows)


def save_exames():
    rows = []
    for e in exames:
        rows.append({
            "nome": e.get("nome", ""),
            "tipo": e.get("tipo", ""),
            "descricao": e.get("descricao", "")
        })
    _write_csv(EXAMES_CSV, ["nome", "tipo", "descricao"], rows)


def save_disponibilidade():
    rows = []
    for d in disponibilidade:
        dh = d.get('data_hora')
        if isinstance(dh, datetime):
            dhs = dh.strftime("%d/%m/%Y %H:%M")
        else:
            dhs = str(dh)
        rows.append({
            "medico": d.get('medico', ''),
            "data_hora": dhs,
            "tipo": d.get('tipo', '')
        })
    _write_csv(DISPO_CSV, ["medico", "data_hora", "tipo"], rows)


def save_agendamentos():
    rows = []
    for a in agendamentos:
        dh = a.get('data_hora')
        if isinstance(dh, datetime):
            dhs = dh.strftime("%d/%m/%Y %H:%M")
        else:
            dhs = str(dh)
        rows.append({
            "paciente": a.get("paciente", ""),
            "medico": a.get("medico", ""),
            "data_hora": dhs,
            "tipo": a.get("tipo", ""),
            "exame": a.get("exame", "")
        })
    _write_csv(AGENDAMENTOS_CSV, ["paciente",
               "medico", "data_hora", "tipo", "exame"], rows)


def save_atendimentos():
    rows = []
    for h in atendimentos:
        dh = h.get('data_hora')
        if isinstance(dh, datetime):
            dhs = dh.strftime("%d/%m/%Y %H:%M")
        else:
            dhs = str(dh)
        rows.append({
            "paciente": h.get("paciente", ""),
            "medico": h.get("medico", ""),
            "data_hora": dhs,
            "descricao": h.get("descricao", "")
        })
    _write_csv(ATENDIMENTOS_CSV, ["paciente",
               "medico", "data_hora", "descricao"], rows)


def save_cadastros_atendimento():
    rows = []
    for registro in cadastros_atendimento:
        dh = registro.get('data_registro')
        if isinstance(dh, datetime):
            dhs = dh.strftime("%d/%m/%Y %H:%M")
        else:
            dhs = str(dh)
        rows.append({
            "paciente_id": registro.get('paciente_id', ''),
            "paciente_nome": registro.get('paciente_nome', ''),
            "cpf": registro.get('cpf', ''),
            "convenio": registro.get('convenio', ''),
            "carteirinha": registro.get('carteirinha', ''),
            "data_registro": dhs
        })
    _write_csv(CADASTROS_ATENDIMENTO_CSV,
               ["paciente_id", "paciente_nome", "cpf", "convenio", "carteirinha", "data_registro"], rows)


def save_convenios():
    rows = []
    for convenio in convenios:
        rows.append({
            "nome": convenio.get('nome', ''),
            "cobertura": convenio.get('cobertura', '')
        })
    _write_csv(CONVENIOS_CSV, ["nome", "cobertura"], rows)


def export_data():
    """Exporta os dados de pacientes, médicos e exames para um arquivo .xlsx.

    O usuário informa o caminho do arquivo .xlsx de destino. Se openpyxl não
    estiver instalado, informa o usuário e cancela a operação.
    """
    # export_data removida: persistência agora é feita automaticamente no arquivo mestre XLSX
    # Use save_master_xlsx() se precisar salvar manualmente.


def save_master_xlsx():
    """Salva todas as abas (pacientes, medicos, exames, agendamentos) no arquivo MASTER_XLSX.

    Se openpyxl não estiver disponível, apenas salva os CSVs (comunica que o XLSX não foi salvo).
    """
    # garantir que os CSVs também existam/estejam atualizados
    save_pacientes()
    save_medicos()
    save_exames()
    save_agendamentos()
    save_atendimentos()
    save_cadastros_atendimento()
    save_convenios()

    if not HAVE_OPENPYXL:
        print("Aviso: openpyxl não disponível — não foi possível salvar a planilha mestre XLSX. Dados salvos em CSV.")
        return

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'pacientes'
        ws.append(['id', 'nome', 'idade', 'telefone'])
        for p in pacientes:
            ws.append([p.get('id', ''), p.get('nome', ''), p.get(
                'idade', ''), p.get('telefone', '')])

        ws2 = wb.create_sheet('medicos')
        ws2.append(['nome', 'especialidade', 'CRM', 'telefone'])
        for m in medicos:
            ws2.append([m.get('nome', ''), m.get('especialidade', ''),
                       m.get('CRM', ''), m.get('telefone', '')])

        ws3 = wb.create_sheet('exames')
        ws3.append(['nome', 'tipo', 'descricao'])
        for e in exames:
            ws3.append([e.get('nome', ''), e.get(
                'tipo', ''), e.get('descricao', '')])

        # agendamentos
        ws4 = wb.create_sheet('agendamentos')
        ws4.append(['paciente', 'medico', 'data_hora', 'tipo', 'exame'])
        for a in agendamentos:
            dh = a.get('data_hora')
            if isinstance(dh, datetime):
                dhs = dh.strftime("%d/%m/%Y %H:%M")
            else:
                dhs = str(dh)
            ws4.append([a.get('paciente', ''), a.get(
                'medico', ''), dhs, a.get('tipo', ''), a.get('exame', '')])

        # disponibilidade
        ws5 = wb.create_sheet('disponibilidade')
        ws5.append(['medico', 'data_hora', 'tipo'])
        for d in disponibilidade:
            ddh = d.get('data_hora')
            if isinstance(ddh, datetime):
                ddhs = ddh.strftime("%d/%m/%Y %H:%M")
            else:
                ddhs = str(ddh)
            ws5.append([d.get('medico', ''), ddhs, d.get('tipo', '')])

        # atendimentos
        ws6 = wb.create_sheet('atendimentos')
        ws6.append(['paciente', 'medico', 'data_hora', 'descricao'])
        for h in atendimentos:
            hdh = h.get('data_hora')
            if isinstance(hdh, datetime):
                hdh_str = hdh.strftime("%d/%m/%Y %H:%M")
            else:
                hdh_str = str(hdh)
            ws6.append([h.get('paciente', ''), h.get(
                'medico', ''), hdh_str, h.get('descricao', '')])

        # horários agregados por médico e data (colunas: medico, data, horarios)
        try:
            from collections import defaultdict
            horarios_map = defaultdict(lambda: defaultdict(list))
            for d in disponibilidade:
                medico_n = (d.get('medico') or '').strip()
                dh = d.get('data_hora')
                if isinstance(dh, str):
                    dhc = validar_data_hora(dh)
                    if dhc:
                        dh = dhc
                if not isinstance(dh, datetime):
                    continue
                data_str = dh.strftime('%d/%m/%Y')
                time_str = dh.strftime('%H:%M')
                horarios_map[medico_n][data_str].append(time_str)

            ws7 = wb.create_sheet('horarios')
            ws7.append(['medico', 'data', 'horarios'])
            for medico_n, datas in sorted(horarios_map.items()):
                for data_str, times in sorted(datas.items()):
                    times_sorted = sorted(times)
                    horarios_joined = "; ".join(times_sorted)
                    ws7.append([medico_n, data_str, horarios_joined])
        except Exception:
            # não fatal: se houver problema na geração da aba 'horarios', prosseguir
            pass

        ws8 = wb.create_sheet('convenios')
        ws8.append(['nome', 'cobertura'])
        for convenio in convenios:
            ws8.append([convenio.get('nome', ''),
                       convenio.get('cobertura', '')])

        ws9 = wb.create_sheet('cadastro_atendimentos')
        ws9.append(['paciente_id', 'paciente_nome', 'cpf',
                    'convenio', 'carteirinha', 'data_registro'])
        for registro in cadastros_atendimento:
            dh_reg = registro.get('data_registro')
            if isinstance(dh_reg, datetime):
                dh_reg_str = dh_reg.strftime("%d/%m/%Y %H:%M")
            else:
                dh_reg_str = str(dh_reg)
            ws9.append([
                registro.get('paciente_id', ''),
                registro.get('paciente_nome', ''),
                registro.get('cpf', ''),
                registro.get('convenio', ''),
                registro.get('carteirinha', ''),
                dh_reg_str
            ])

        os.makedirs(os.path.dirname(MASTER_XLSX), exist_ok=True)
        wb.save(MASTER_XLSX)
        # mensagem simples
        # print(f"Planilha mestre salva em: {MASTER_XLSX}")
    except Exception as e:
        print(f"Erro ao salvar planilha mestre XLSX: {e}")


def load_master_xlsx():
    """Carrega pacientes, medicos, exames e agendamentos do MASTER_XLSX se disponível.

    Retorna True se o carregamento foi feito a partir do XLSX, False caso contrário.
    """
    if not HAVE_OPENPYXL:
        return False
    if not os.path.exists(MASTER_XLSX):
        return False

    try:
        wb = load_workbook(MASTER_XLSX)
    except Exception:
        return False

    # pacientes
    pacientes.clear()
    if 'pacientes' in wb.sheetnames:
        ws = wb['pacientes']
        rows = list(ws.values)
        if rows:
            headers = [str(h).strip().lower() for h in rows[0]]
            header_com_id = ['id', 'nome', 'idade', 'telefone']
            header_sem_id = ['nome', 'idade', 'telefone']
            if headers == header_com_id:
                idx_id, idx_nome, idx_idade, idx_tel = 0, 1, 2, 3
            elif headers == header_sem_id:
                idx_id, idx_nome, idx_idade, idx_tel = None, 0, 1, 2
            else:
                print(
                    f"Aviso: cabeçalho inesperado em 'pacientes' na planilha mestre. Ignorando XLSX.")
                return False

            for r in rows[1:]:
                nome = r[idx_nome] if len(r) > idx_nome else ''
                idade = r[idx_idade] if len(r) > idx_idade else None
                telefone = r[idx_tel] if len(r) > idx_tel else ''
                paciente_id = ''
                if idx_id is not None and len(r) > idx_id:
                    paciente_id = r[idx_id]
                try:
                    idade_val = int(idade) if idade not in (None, '') else None
                except Exception:
                    idade_val = None
                pacientes.append({
                    'id': str(paciente_id or '').strip(),
                    'nome': nome or '',
                    'idade': idade_val,
                    'telefone': telefone or ''
                })

    # medicos
    medicos.clear()
    if 'medicos' in wb.sheetnames:
        ws = wb['medicos']
        rows = list(ws.values)
        if rows:
            headers = [str(h).strip().lower() for h in rows[0]]
            expected = ['nome', 'especialidade', 'crm', 'telefone']
            if headers != expected:
                print(
                    f"Aviso: cabeçalho inesperado em 'medicos' na planilha mestre. Ignorando XLSX.")
                return False
            for r in rows[1:]:
                nome = r[0] if len(r) > 0 else ''
                especialidade = r[1] if len(r) > 1 else ''
                crm = r[2] if len(r) > 2 else ''
                telefone = r[3] if len(r) > 3 else ''
                medicos.append({'nome': nome or '', 'especialidade': especialidade or '',
                               'CRM': crm or '', 'telefone': telefone or ''})

    # exames
    exames.clear()
    if 'exames' in wb.sheetnames:
        ws = wb['exames']
        rows = list(ws.values)
        if rows:
            headers = [str(h).strip().lower() for h in rows[0]]
            expected = ['nome', 'tipo', 'descricao']
            if headers != expected:
                print(
                    f"Aviso: cabeçalho inesperado em 'exames' na planilha mestre. Ignorando XLSX.")
                return False
            for r in rows[1:]:
                nome = r[0] if len(r) > 0 else ''
                tipo = r[1] if len(r) > 1 else ''
                descricao = r[2] if len(r) > 2 else ''
                exames.append(
                    {'nome': nome or '', 'tipo': tipo or '', 'descricao': descricao or ''})

    # agendamentos
    agendamentos.clear()
    if 'agendamentos' in wb.sheetnames:
        ws = wb['agendamentos']
        rows = list(ws.values)
        if rows:
            headers = [str(h).strip().lower() for h in rows[0]]
            expected_new = ['paciente', 'medico', 'data_hora', 'tipo', 'exame']
            expected_old = ['paciente', 'medico', 'data_hora', 'tipo']
            if headers not in (expected_new, expected_old):
                print(
                    "Aviso: cabeçalho inesperado em 'agendamentos' na planilha mestre. Ignorando XLSX.")
                return False
            for r in rows[1:]:
                paciente = r[0] if len(r) > 0 else ''
                medico = r[1] if len(r) > 1 else ''
                data_hora = r[2] if len(r) > 2 else ''
                tipo = r[3] if len(r) > 3 else ''
                exame = r[4] if len(r) > 4 else ''
                agendamentos.append({'paciente': paciente or '', 'medico': medico or '',
                                    'data_hora': data_hora or '', 'tipo': tipo or '', 'exame': exame or ''})

    # atendimentos
    atendimentos.clear()
    if 'atendimentos' in wb.sheetnames:
        ws = wb['atendimentos']
        rows = list(ws.values)
        if rows:
            for r in rows[1:]:
                paciente = r[0] if len(r) > 0 else ''
                medico = r[1] if len(r) > 1 else ''
                data_hora = r[2] if len(r) > 2 else ''
                descricao = r[3] if len(r) > 3 else ''
                dh = data_hora
                if isinstance(dh, str):
                    conv = validar_data_hora(dh)
                    if conv:
                        dh = conv
                atendimentos.append({
                    'paciente': paciente or '',
                    'medico': medico or '',
                    'data_hora': dh or '',
                    'descricao': descricao or ''
                })

    cadastros_atendimento.clear()
    if 'cadastro_atendimentos' in wb.sheetnames:
        ws = wb['cadastro_atendimentos']
        rows = list(ws.values)
        if rows:
            headers = [str(h).strip().lower() for h in rows[0]]
            expected = ['paciente_id', 'paciente_nome', 'cpf',
                        'convenio', 'carteirinha', 'data_registro']
            if headers == expected:
                for r in rows[1:]:
                    pid = r[0] if len(r) > 0 else ''
                    nome = r[1] if len(r) > 1 else ''
                    cpf = r[2] if len(r) > 2 else ''
                    convenio = r[3] if len(r) > 3 else ''
                    carteirinha = r[4] if len(r) > 4 else ''
                    data_reg = r[5] if len(r) > 5 else ''
                    dh = data_reg
                    if isinstance(dh, str):
                        conv = validar_data_hora(dh)
                        if conv:
                            dh = conv
                    cadastros_atendimento.append({
                        'paciente_id': str(pid or '').strip(),
                        'paciente_nome': nome or '',
                        'cpf': cpf or '',
                        'convenio': convenio or '',
                        'carteirinha': carteirinha or '',
                        'data_registro': dh or ''
                    })

    convenios.clear()
    if 'convenios' in wb.sheetnames:
        ws = wb['convenios']
        rows = list(ws.values)
        if rows:
            headers = [str(h).strip().lower() for h in rows[0]]
            if headers == ['nome', 'cobertura']:
                for r in rows[1:]:
                    nome = r[0] if len(r) > 0 else ''
                    cobertura = r[1] if len(r) > 1 else ''
                    convenios.append({
                        'nome': nome or '',
                        'cobertura': cobertura or ''
                    })

    # disponibilidade
    disponibilidade.clear()
    if 'disponibilidade' in wb.sheetnames:
        ws = wb['disponibilidade']
        rows = list(ws.values)
        if rows:
            for r in rows[1:]:
                medico = r[0] if len(r) > 0 else ''
                data_hora = r[1] if len(r) > 1 else ''
                tipo = r[2] if len(r) > 2 else ''
                dh = data_hora
                if isinstance(dh, str):
                    conv = validar_data_hora(dh)
                    if conv:
                        dh = conv
                disponibilidade.append(
                    {'medico': medico or '', 'data_hora': dh or '', 'tipo': (tipo or '').lower()})

    return True


def import_data():
    """Importa dados de um arquivo .xlsx contendo as abas 'pacientes','medicos' e 'exames'.

    Se openpyxl não estiver disponível, avisa o usuário e cancela.
    """
    # import_data removida: o sistema carrega automaticamente a planilha mestre na inicialização
    # e salva nela sempre que há alterações. Se precisar importar manualmente, use load_master_xlsx().


def load_data():
    # carregar pacientes
    for row in _read_csv(PACIENTES_CSV):
        try:
            idade = int(row.get('idade', '') or 0) if row.get(
                'idade', '') else None
        except ValueError:
            idade = None
        pid = str(row.get('id', '') or '').strip()
        pacientes.append({
            "id": pid,
            "nome": row.get('nome', ''),
            "idade": idade,
            "telefone": row.get('telefone', '')
        })

    # carregar medicos
    for row in _read_csv(MEDICOS_CSV):
        medicos.append({
            "nome": row.get('nome', ''),
            "especialidade": row.get('especialidade', ''),
            "CRM": row.get('CRM', ''),
            "telefone": row.get('telefone', '')
        })

    # carregar exames
    for row in _read_csv(EXAMES_CSV):
        exames.append({
            "nome": row.get('nome', ''),
            "tipo": row.get('tipo', ''),
            "descricao": row.get('descricao', '')
        })

    # carregar disponibilidade
    for row in _read_csv(DISPO_CSV):
        dh = row.get('data_hora', '')
        conv = None
        if dh:
            try:
                conv = validar_data_hora(dh)
            except Exception:
                conv = None
        disponibilidade.append({
            'medico': row.get('medico', ''),
            'data_hora': conv or dh,
            'tipo': (row.get('tipo', '') or '').lower()
        })

    # carregar agendamentos (data_hora fica string por enquanto; migrate irá converter)
    for row in _read_csv(AGENDAMENTOS_CSV):
        agendamentos.append({
            "paciente": row.get('paciente', ''),
            "medico": row.get('medico', ''),
            "data_hora": row.get('data_hora', ''),
            "tipo": row.get('tipo', ''),
            "exame": row.get('exame', '')
        })

    # carregar atendimentos (data_hora fica string por enquanto; migrate irá converter)
    for row in _read_csv(ATENDIMENTOS_CSV):
        atendimentos.append({
            "paciente": row.get('paciente', ''),
            "medico": row.get('medico', ''),
            "data_hora": row.get('data_hora', ''),
            "descricao": row.get('descricao', '')
        })

    convenios.clear()
    for row in _read_csv(CONVENIOS_CSV):
        convenios.append({
            "nome": row.get('nome', ''),
            "cobertura": row.get('cobertura', '')
        })

    for row in _read_csv(CADASTROS_ATENDIMENTO_CSV):
        cadastros_atendimento.append({
            "paciente_id": row.get('paciente_id', ''),
            "paciente_nome": row.get('paciente_nome', ''),
            "cpf": row.get('cpf', ''),
            "convenio": row.get('convenio', ''),
            "carteirinha": row.get('carteirinha', ''),
            "data_registro": row.get('data_registro', '')
        })


def validar_data_hora(data_hora):
    try:
        return datetime.strptime(data_hora, "%d/%m/%Y %H:%M")
    except ValueError:
        return None


def horario_ja_cadastrado(medico_nome, tipo, data_hora):
    alvo_nome = (medico_nome or "").strip().lower()
    alvo_tipo = (tipo or "").strip().lower()
    for slot in disponibilidade:
        nome_slot = (slot.get('medico', '') or "").strip().lower()
        tipo_slot = (slot.get('tipo', '') or "").strip().lower()
        if nome_slot != alvo_nome:
            continue
        if tipo_slot != alvo_tipo and tipo_slot != 'ambos' and alvo_tipo != 'ambos':
            continue
        slot_dh = slot.get('data_hora')
        if isinstance(slot_dh, datetime):
            if slot_dh == data_hora:
                return True
        elif isinstance(slot_dh, str):
            conv = validar_data_hora(slot_dh)
            if conv and conv == data_hora:
                return True
    return False


# carregar dados persistidos: preferir a planilha mestre XLSX quando existir
loaded_from_xlsx = load_master_xlsx()
if not loaded_from_xlsx:
    load_data()
sincronizar_ids_pacientes()
# migrar agendamentos (converte strings para datetime)
migrate_agendamentos()
migrate_atendimentos()


def cadastrar_paciente():
    try:
        nome = input("Nome do paciente: ").strip()
        if not validar_nome(nome):
            print("Erro: o nome deve conter apenas letras e espaços.")
            return
        idade = int(input("Idade: "))
        telefone = input("Telefone: ").strip()
        telefone = format_telefone(telefone)
        sincronizar_ids_pacientes()
        novo_id = gerar_id_paciente()
        paciente = {"id": novo_id, "nome": nome,
                    "idade": idade, "telefone": telefone}
        pacientes.append(paciente)
        save_pacientes()
        # atualizar planilha mestre imediatamente
        save_master_xlsx()
        print(f"Paciente cadastrado com sucesso! ID: {novo_id}\n")
    except ValueError:
        print("Erro: idade deve ser um número inteiro.\n")


def cadastrar_medico():
    try:
        nome = input("Nome do médico: ").strip()
        if not validar_nome(nome):
            print("Erro: o nome deve conter apenas letras e espaços.")
            return
        especialidade = input("Especialidade: ").strip()
        telefone = input("Telefone: ").strip()
        telefone = format_telefone(telefone)
        crm = input("CRM: ").strip()
        medico = {"nome": nome, "especialidade": especialidade,
                  "CRM": crm, "telefone": telefone}
        medicos.append(medico)
        save_medicos()
        # atualizar planilha mestre imediatamente
        save_master_xlsx()
        print("Médico cadastrado com sucesso!\n")
    except Exception as e:
        print(f"Erro ao cadastrar médico: {e}\n")


def cadastrar_exame():
    try:
        nome = input("Nome do exame: ").strip()
        tipo = input("Tipo do exame: ").strip()
        descricao = input("Descrição: ").strip()
        if not descricao:
            print("Erro: a descrição não pode ficar em branco.")
            return
        exame = {"nome": nome, "tipo": tipo, "descricao": descricao}
        exames.append(exame)
        save_exames()
        # atualizar planilha mestre imediatamente
        save_master_xlsx()
        print("Exame cadastrado com sucesso!\n")
    except Exception as e:
        print(f"Erro ao cadastrar exame: {e}\n")


def cadastrar_convenio():
    try:
        nome = input("Nome do convênio: ").strip()
        if not nome:
            print("Erro: o nome do convênio não pode ficar em branco.")
            return
        cobertura = input("Cobertura ou observações: ").strip()
        convenios.append({"nome": nome, "cobertura": cobertura})
        save_convenios()
        save_master_xlsx()
        print("Convênio cadastrado com sucesso!\n")
    except Exception as e:
        print(f"Erro ao cadastrar convênio: {e}\n")


def listar_convenios():
    print("\n--- Convênios cadastrados ---")
    if not convenios:
        print("Nenhum convênio cadastrado.\n")
        return
    for idx, convenio in enumerate(convenios, start=1):
        cobertura = convenio.get('cobertura')
        cobertura_info = f" | Cobertura: {cobertura}" if cobertura else ""
        print(f"{idx}. {convenio.get('nome', '')}{cobertura_info}")
    print()


def cadastrar_disponibilidade():
    try:
        if not medicos:
            print("Erro: não há médicos cadastrados.")
            return

        medico = None
        while medico is None:
            termo = input(
                "Nome (ou parte do nome) do médico: ").strip()
            if not termo:
                print("Informe ao menos uma letra para buscar o médico.")
                continue

            termo_norm = normalizar_texto(termo)
            candidatos = [m for m in medicos if termo_norm
                          in normalizar_texto(m.get('nome', '').strip())]

            if not candidatos:
                print("Nenhum médico encontrado com esse termo.")
                while True:
                    print("1. Listar médicos cadastrados")
                    print("0. Voltar ao menu")
                    acao = input("Escolha uma opção: ").strip()
                    if acao == "0":
                        print("Voltando ao menu anterior.\n")
                        return
                    if acao == "1":
                        if not medicos:
                            print("Não há médicos cadastrados.")
                            return
                        print("\n--- Médicos Cadastrados ---")
                        for idx, cand in enumerate(medicos, start=1):
                            espec = cand.get('especialidade')
                            espec_str = f" | Especialidade: {espec}" if espec else ""
                            print(f"{idx}. {cand.get('nome', '')}{espec_str}")
                        print("0. Voltar")
                        print("M. Voltar ao menu anterior")
                        escolha_lista = input(
                            "Escolha o número do médico: ").strip()
                        if escolha_lista.lower() == 'm':
                            print("Voltando ao menu anterior.\n")
                            return
                        if escolha_lista.isdigit():
                            escolha_int = int(escolha_lista)
                            if escolha_int == 0:
                                print()
                                break
                            if 1 <= escolha_int <= len(medicos):
                                medico = medicos[escolha_int - 1]
                                print()
                                break
                        print("Seleção inválida. Tente novamente.")
                        continue
                    print("Opção inválida. Tente novamente.")
                if medico:
                    break
                continue

            while True:
                print("Médicos encontrados:")
                for idx, cand in enumerate(candidatos, start=1):
                    espec = cand.get('especialidade')
                    espec_str = f" | Especialidade: {espec}" if espec else ""
                    print(f"{idx}. {cand.get('nome', '')}{espec_str}")
                print("V. Pesquisar novamente")
                print("0. Voltar ao menu anterior")
                escolha_cand = input(
                    "Escolha o número do médico: ").strip().lower()
                if escolha_cand == '0':
                    print("Voltando ao menu anterior.\n")
                    return
                if escolha_cand == 'v':
                    medico = None
                    break
                if escolha_cand.isdigit():
                    escolha_int = int(escolha_cand)
                    if 1 <= escolha_int <= len(candidatos):
                        medico = candidatos[escolha_int - 1]
                        break
                print("Seleção inválida. Tente novamente.")
            if medico:
                break
            continue

        medico_nome = medico.get('nome', '')

        tipo = None
        while tipo is None:
            print("\nSelecione o tipo de horário permitido:")
            print("1. Exame")
            print("2. Consulta")
            print("0. Voltar ao menu anterior")
            escolha_tipo = input("Escolha uma opção: ").strip()
            if escolha_tipo == "0":
                print("Voltando ao menu anterior.\n")
                return
            if escolha_tipo == "1":
                tipo = "exame"
            elif escolha_tipo == "2":
                tipo = "consulta"
            else:
                print("Opção inválida. Tente novamente.")

        print("Digite as datas/horários disponíveis para este médico. Deixe em branco para terminar.")
        while True:
            data_str = input(
                "Data (DD/MM/AAAA) [enter para sair]: ").strip()
            if not data_str:
                break
            try:
                data_base = datetime.strptime(data_str, "%d/%m/%Y")
            except ValueError:
                print("Formato de data inválido. Tente novamente.")
                continue

            while True:
                print("\nComo deseja informar os horários para esta data?")
                print("1. Inserir horários manualmente")
                print("2. Gerar horários automaticamente (intervalo fixo)")
                print("0. Cancelar esta data")
                escolha_horario = input("Escolha uma opção: ").strip()

                if escolha_horario == "0":
                    print("Data descartada.\n")
                    break

                if escolha_horario == "1":
                    while True:
                        hora_str = input(
                            "Horário (HH:MM) [enter para terminar esta data]: ").strip()
                        if not hora_str:
                            break
                        dt_composta = f"{data_str} {hora_str}"
                        dh = validar_data_hora(dt_composta)
                        if not dh:
                            print("Formato de horário inválido. Tente novamente.")
                            continue
                        if horario_ja_cadastrado(medico_nome, tipo, dh):
                            print(
                                "Aviso: horário já cadastrado para este médico e tipo. Ignorando duplicata.")
                            continue
                        disponibilidade.append(
                            {'medico': medico_nome, 'data_hora': dh, 'tipo': tipo})
                    if escolha_horario == "1":
                        break

                elif escolha_horario == "2":
                    inicio_str = input("Horário inicial (HH:MM): ").strip()
                    fim_str = input("Horário final (HH:MM): ").strip()
                    intervalo_str = input(
                        "Intervalo entre horários em minutos: ").strip()
                    if not inicio_str or not fim_str or not intervalo_str:
                        print(
                            "Preencha todos os campos para gerar os horários automaticamente.")
                        continue
                    try:
                        inicio_dt = datetime.strptime(
                            f"{data_str} {inicio_str}", "%d/%m/%Y %H:%M")
                        fim_dt = datetime.strptime(
                            f"{data_str} {fim_str}", "%d/%m/%Y %H:%M")
                        intervalo_min = int(intervalo_str)
                    except ValueError:
                        print(
                            "Dados inválidos. Verifique horários e intervalo informados.")
                        continue
                    if intervalo_min <= 0:
                        print("O intervalo deve ser um número positivo.")
                        continue
                    if fim_dt <= inicio_dt:
                        print("O horário final deve ser posterior ao inicial.")
                        continue

                    horarios_adicionados = 0
                    atual = inicio_dt
                    while atual <= fim_dt:
                        if not horario_ja_cadastrado(medico_nome, tipo, atual):
                            disponibilidade.append(
                                {'medico': medico_nome, 'data_hora': atual, 'tipo': tipo})
                            horarios_adicionados += 1
                        else:
                            print(
                                f"Aviso: {atual.strftime('%H:%M')} já estava cadastrado e foi ignorado.")
                        atual += timedelta(minutes=intervalo_min)
                    if horarios_adicionados:
                        print(
                            f"{horarios_adicionados} horário(s) adicionados para {data_str}.")
                    else:
                        print("Nenhum horário novo foi adicionado para esta data.")
                    break

                else:
                    print("Opção inválida. Tente novamente.")
                    continue
            # prosseguir para a próxima data
        save_disponibilidade()
        save_master_xlsx()
        print("Disponibilidade cadastrada/atualizada para o médico.\n")
    except Exception as e:
        print(f"Erro ao cadastrar disponibilidade: {e}\n")


def registrar_atendimento(paciente_preselecionado=None, medico_autenticado=None, mostrar_agenda=True):
    if not medicos:
        print("Erro: não há médicos cadastrados.")
        return
    if not pacientes:
        print("Erro: não há pacientes cadastrados.")
        return

    sincronizar_ids_pacientes()

    # seleção do médico responsável
    medico = medico_autenticado
    while medico is None:
        termo_medico = input(
            "Nome (ou parte do nome) do médico responsável: ").strip()
        if not termo_medico:
            print("Informe ao menos uma letra para buscar o médico.")
            continue
        termo_norm = normalizar_texto(termo_medico)
        candidatos = [m for m in medicos if termo_norm in normalizar_texto(
            m.get('nome', '').strip())]
        if not candidatos:
            print("Nenhum médico encontrado com esse termo.")
            while True:
                print("1. Listar médicos cadastrados")
                print("0. Voltar ao menu anterior")
                acao = input("Escolha uma opção: ").strip()
                if acao == "0":
                    print()
                    return
                if acao == "1":
                    print("\n--- Médicos Cadastrados ---")
                    for idx, cand in enumerate(medicos, start=1):
                        espec = cand.get('especialidade')
                        espec_str = f" | Especialidade: {espec}" if espec else ""
                        print(f"{idx}. {cand.get('nome', '')}{espec_str}")
                    print("0. Voltar")
                    escolha_lista = input(
                        "Escolha o número do médico: ").strip()
                    if escolha_lista.isdigit():
                        escolha_int = int(escolha_lista)
                        if escolha_int == 0:
                            print()
                            break
                        if 1 <= escolha_int <= len(medicos):
                            medico = medicos[escolha_int - 1]
                            print()
                            break
                    print("Seleção inválida. Tente novamente.")
                    continue
                print("Opção inválida. Tente novamente.")
            continue

        while True:
            print("Médicos encontrados:")
            for idx, cand in enumerate(candidatos, start=1):
                espec = cand.get('especialidade')
                espec_str = f" | Especialidade: {espec}" if espec else ""
                print(f"{idx}. {cand.get('nome', '')}{espec_str}")
            print("0. Pesquisar novamente")
            print("M. Voltar ao menu anterior")
            escolha = input("Escolha o número do médico: ").strip()
            if escolha.lower() == 'm':
                print()
                return
            if escolha.isdigit():
                escolha_int = int(escolha)
                if escolha_int == 0:
                    medico = None
                    break
                if 1 <= escolha_int <= len(candidatos):
                    medico = candidatos[escolha_int - 1]
                    break
            print("Seleção inválida. Tente novamente.")
        if medico:
            break

    medico_nome = medico.get('nome', '').strip()

    # seleção da data
    data_base = None
    while data_base is None:
        data_str = input("Data do atendimento (DD/MM/AAAA): ").strip()
        if not data_str:
            print("Informe uma data válida.")
            continue
        try:
            data_base = datetime.strptime(data_str, "%d/%m/%Y")
        except ValueError:
            print("Formato de data inválido. Tente novamente.")
            continue

    # agenda do médico na data selecionada
    agenda_dia = []
    for ag in agendamentos:
        if normalizar_texto(ag.get('medico', '').strip()) != normalizar_texto(medico_nome):
            continue
        dh = ag.get('data_hora')
        if isinstance(dh, str):
            dh = validar_data_hora(dh)
        if isinstance(dh, datetime) and dh.date() == data_base.date():
            agenda_dia.append(ag)

    agenda_dia = sorted(agenda_dia, key=lambda registro: registro.get('data_hora')
                        if isinstance(registro.get('data_hora'), datetime)
                        else validar_data_hora(registro.get('data_hora') or '') or datetime.max)

    if mostrar_agenda:
        if agenda_dia:
            print(
                f"\n--- Agenda de {medico_nome} em {data_base.strftime('%d/%m/%Y')} ---")
            for idx, registro in enumerate(agenda_dia, start=1):
                dh = registro.get('data_hora')
                if isinstance(dh, datetime):
                    dh_str = dh.strftime("%H:%M")
                else:
                    conv = validar_data_hora(
                        dh) if isinstance(dh, str) else None
                    dh_str = conv.strftime("%H:%M") if conv else "--:--"
                tipo = registro.get('tipo') or 'N/A'
                exame = registro.get('exame')
                exame_str = f" | Exame: {exame}" if exame else ""
                print(
                    f"{idx}. {dh_str} | Paciente: {registro.get('paciente', '')} | Tipo: {tipo}{exame_str}")
        else:
            print(
                f"Nenhum agendamento encontrado para {medico_nome} nesta data.")

    paciente = None
    if paciente_preselecionado:
        paciente_existente = localizar_paciente_por_identificadores(
            paciente_preselecionado.get('id'), paciente_preselecionado.get('nome'))
        if paciente_existente:
            paciente = paciente_existente
            print(
                f"\nPaciente pré-selecionado para o atendimento: {paciente.get('nome', '')}")
        else:
            print(
                "\nAviso: paciente pré-selecionado não encontrado nos cadastros. Selecione manualmente.")

    data_hora_atendimento = None

    while paciente is None:
        if mostrar_agenda and agenda_dia:
            print(
                "\nSelecione um paciente da agenda acima ou digite B para buscar outro paciente.")
            escolha_agenda = input("Número do paciente (ou B): ").strip()
            if escolha_agenda.lower() == 'b':
                pass  # cairá na busca manual abaixo
            elif escolha_agenda.isdigit():
                escolha_int = int(escolha_agenda)
                if 1 <= escolha_int <= len(agenda_dia):
                    agendamento_escolhido = agenda_dia[escolha_int - 1]
                    nome_paciente = agendamento_escolhido.get(
                        'paciente', '').strip()
                    paciente = next((p for p in pacientes if normalizar_texto(p.get('nome', '').strip())
                                     == normalizar_texto(nome_paciente)), None)
                    if paciente is None:
                        print(
                            "Paciente selecionado não está cadastrado. Selecione outro.")
                        continue
                    dh = agendamento_escolhido.get('data_hora')
                    if isinstance(dh, str):
                        dh = validar_data_hora(dh)
                    if isinstance(dh, datetime):
                        data_hora_atendimento = dh
                    break
                else:
                    print("Seleção inválida. Tente novamente.")
                    continue

        # busca manual por paciente
        termo_paciente = input("Nome (ou parte do nome) do paciente: ").strip()
        if not termo_paciente:
            print("Informe ao menos uma letra para buscar o paciente.")
            continue
        termo_pac_norm = normalizar_texto(termo_paciente)
        candidatos_pac = [p for p in pacientes if termo_pac_norm
                          in normalizar_texto(p.get('nome', '').strip())]
        if not candidatos_pac:
            print("Nenhum paciente encontrado com esse termo.")
            while True:
                print("1. Listar pacientes cadastrados")
                print("0. Voltar ao menu anterior")
                acao = input("Escolha uma opção: ").strip()
                if acao == "0":
                    print()
                    return
                if acao == "1":
                    print("\n--- Pacientes Cadastrados ---")
                    for idx, cand in enumerate(pacientes, start=1):
                        idade_info = cand.get('idade')
                        idade_str = f" | Idade: {idade_info}" if isinstance(
                            idade_info, (int, float)) else ""
                        pid = obter_id_legivel(cand)
                        print(
                            f"{idx}. [{pid}] {cand.get('nome', '')}{idade_str}")
                    print("0. Voltar")
                    escolha = input("Escolha o número do paciente: ").strip()
                    if escolha.isdigit():
                        escolha_int = int(escolha)
                        if escolha_int == 0:
                            print()
                            break
                        if 1 <= escolha_int <= len(pacientes):
                            paciente = pacientes[escolha_int - 1]
                            print()
                            break
                    print("Seleção inválida. Tente novamente.")
                    continue
                print("Opção inválida. Tente novamente.")
            if paciente:
                break
            continue

        if len(candidatos_pac) == 1:
            paciente = candidatos_pac[0]
        else:
            print("Pacientes encontrados:")
            for idx, cand in enumerate(candidatos_pac, start=1):
                idade_info = cand.get('idade')
                idade_str = f" | Idade: {idade_info}" if isinstance(
                    idade_info, (int, float)) else ""
                pid = obter_id_legivel(cand)
                print(f"{idx}. [{pid}] {cand.get('nome', '')}{idade_str}")
            print("0. Pesquisar novamente")
            print("M. Voltar ao menu anterior")
            escolha = input("Escolha o número do paciente: ").strip()
            if escolha.lower() == 'm':
                print()
                return
            if escolha.isdigit():
                escolha_int = int(escolha)
                if escolha_int == 0:
                    paciente = None
                    continue
                if 1 <= escolha_int <= len(candidatos_pac):
                    paciente = candidatos_pac[escolha_int - 1]
                else:
                    print("Seleção inválida. Tente novamente.")
                    paciente = None
                    continue
            else:
                print("Seleção inválida. Tente novamente.")
                paciente = None
                continue

    if data_hora_atendimento is None:
        while True:
            hora_str = input("Horário do atendimento (HH:MM): ").strip()
            if not hora_str:
                print("Informe um horário válido ou retorne ao menu.")
                continue
            try:
                hora_dt = datetime.strptime(hora_str, "%H:%M")
            except ValueError:
                print("Formato de horário inválido. Tente novamente.")
                continue
            data_hora_atendimento = datetime.combine(
                data_base.date(), hora_dt.time())
            break

    descricao = input("Resumo do atendimento: ").strip()
    if not descricao:
        print("Erro: a descrição do atendimento não pode ficar em branco.")
        return

    atendimentos.append({
        'paciente': paciente.get('nome', ''),
        'medico': medico_nome,
        'data_hora': data_hora_atendimento,
        'descricao': descricao
    })
    save_atendimentos()
    save_master_xlsx()
    print("Atendimento registrado com sucesso!\n")


def cadastrar_atendimento_paciente():
    if not pacientes:
        print("Erro: não há pacientes cadastrados.")
        return

    paciente = selecionar_paciente_interativo()
    if not paciente:
        print("Cadastro de atendimento cancelado: paciente não localizado.\n")
        return

    pid_legivel = obter_id_legivel(paciente)
    print(f"Paciente localizado: [{pid_legivel}] {paciente.get('nome', '')}\n")

    tentativas_cpf = 0
    while True:
        cpf_input = input("CPF do paciente (somente números): ").strip()
        cpf_digits = ''.join(ch for ch in cpf_input if ch.isdigit())
        if len(cpf_digits) == 11:
            cpf_formatado = formatar_cpf(cpf_digits)
            break
        print("CPF inválido. Informe exatamente 11 dígitos.")
        tentativas_cpf += 1
        if tentativas_cpf >= 2:
            while True:
                repetir = input(
                    "Deseja tentar novamente? (S/N): ").strip().lower()
                if repetir in ("s", "sim"):
                    tentativas_cpf = 0
                    break
                if repetir in ("n", "nao", "não"):
                    print("Voltando ao menu anterior.\n")
                    return
                print("Opção inválida. Digite S para sim ou N para não.")

    convenio = None
    while convenio is None:
        print("\nSelecione o convênio:")
        print("1. Particular")
        if convenios:
            for idx, info in enumerate(convenios, start=2):
                print(f"{idx}. {info.get('nome', '')}")
        else:
            print("(Nenhum convênio cadastrado; selecione 1 para Particular)")
        print("0. Cancelar cadastro")
        escolha_convenio = input("Escolha uma opção: ").strip()
        if escolha_convenio == '0':
            print("Cadastro de atendimento cancelado.\n")
            return
        if escolha_convenio == '1':
            convenio = 'Particular'
            break
        if convenios and escolha_convenio.isdigit():
            escolha_int = int(escolha_convenio)
            offset = escolha_int - 2
            if 0 <= offset < len(convenios):
                convenio = convenios[offset].get('nome', '') or 'Convênio'
                break
        print("Seleção inválida. Tente novamente.")

    while True:
        carteirinha = input("Número da carteirinha do convênio: ").strip()
        if carteirinha:
            break
        print("O número da carteirinha não pode ficar em branco.")

    data_registro = datetime.now()
    if paciente_tem_agendamento_na_data(paciente.get('nome', ''), data_registro.date()):
        print("\nPaciente possui agendamento para esta data. Rodando avaliação de controle de acesso...\n")
        avaliar_controle_acesso(tem_agendamento=True)
    else:
        listar_especialidades_disponiveis_para_data(data_registro)
        print("Iniciando fluxo de agendamento a partir deste cadastro...\n")
        agendamentos_antes = len(agendamentos)
        agendar(referencia_temporal=data_registro,
                paciente_preselecionado=paciente)
        if len(agendamentos) > agendamentos_antes:
            avaliar_controle_acesso(tem_agendamento=True)

    registro = {
        'paciente_id': paciente.get('id', ''),
        'paciente_nome': paciente.get('nome', ''),
        'cpf': cpf_formatado,
        'convenio': convenio,
        'carteirinha': carteirinha,
        'data_registro': data_registro
    }

    cadastros_atendimento.append(registro)
    save_cadastros_atendimento()
    save_master_xlsx()
    print("Cadastro de atendimento registrado com sucesso!\n")


def buscar_cadastros_atendimento_por_nome():
    if not cadastros_atendimento:
        print("Não há cadastros de atendimento registrados.")
        return

    while True:
        termo = input(
            "Nome (ou parte do nome) do paciente para busca (0 para voltar): ").strip()
        if termo == '0':
            print()
            return
        if not termo:
            print("Informe ao menos uma letra para buscar.")
            continue

        termo_norm = normalizar_texto(termo)
        resultados = [registro for registro in cadastros_atendimento
                      if termo_norm in normalizar_texto(registro.get('paciente_nome', ''))]

        if not resultados:
            print("Nenhum cadastro encontrado para esse termo.")
            continue

        print("\n--- Cadastros encontrados ---")
        for idx, registro in enumerate(resultados, start=1):
            data_reg = registro.get('data_registro')
            if isinstance(data_reg, datetime):
                data_str = data_reg.strftime("%d/%m/%Y %H:%M")
            else:
                data_str = str(data_reg or '')
            print(
                f"{idx}. [{registro.get('paciente_id', '')}] {registro.get('paciente_nome', '')} | CPF: {registro.get('cpf', '')} | Convênio: {registro.get('convenio', '')} | Carteirinha: {registro.get('carteirinha', '')} | Registro: {data_str}")
        print()
        return


def _input_bool(pergunta):
    while True:
        resposta = input(f"{pergunta} (S/N): ").strip().lower()
        if resposta in ("s", "sim"):
            return True
        if resposta in ("n", "nao", "não"):
            return False
        print("Opção inválida. Digite S para sim ou N para não.")


def avaliar_controle_acesso(tem_agendamento=None):
    print("\n--- Controle de Acesso ---")
    if tem_agendamento is None:
        tem_agendamento = _input_bool("A. Paciente tem agendamento?")
    documentos_ok = _input_bool("B. Documentos estão OK?")
    medico_disponivel = _input_bool("C. Médico disponível?")
    pagamentos_em_dia = _input_bool("D. Pagamentos em dia?")

    consulta_normal = avaliar_consulta_normal(
        tem_agendamento, documentos_ok, medico_disponivel, pagamentos_em_dia)
    emergencia = avaliar_emergencia(
        documentos_ok, medico_disponivel, pagamentos_em_dia)

    print("\nResultado das regras de acesso:")
    print(f"Consulta Normal: {'APROVADA' if consulta_normal else 'NEGADA'}")
    print(f"Emergência: {'APROVADA' if emergencia else 'NEGADA'}\n")


def visualizar_historico_paciente():
    if not pacientes:
        print("Erro: não há pacientes cadastrados.")
        return

    sincronizar_ids_pacientes()

    paciente = None
    while paciente is None:
        termo = input("Nome (ou parte do nome) do paciente: ").strip()
        if not termo:
            print("Informe ao menos uma letra para buscar o paciente.")
            continue

        termo_norm = normalizar_texto(termo)
        candidatos = [p for p in pacientes if termo_norm
                      in normalizar_texto(p.get('nome', '').strip())]
        if not candidatos:
            print("Nenhum paciente encontrado com esse termo.")
            while True:
                print("1. Listar pacientes cadastrados")
                print("0. Voltar ao menu principal")
                acao = input("Escolha uma opção: ").strip()
                if acao == "0":
                    print("Voltando ao menu principal.\n")
                    return
                if acao == "1":
                    if not pacientes:
                        print("Não há pacientes cadastrados.")
                        return
                    print("\n--- Pacientes Cadastrados ---")
                    for idx, cand in enumerate(pacientes, start=1):
                        idade_info = cand.get('idade')
                        idade_str = f" | Idade: {idade_info}" if isinstance(
                            idade_info, (int, float)) else ""
                        pid = obter_id_legivel(cand)
                        print(
                            f"{idx}. [{pid}] {cand.get('nome', '')}{idade_str}")
                    print("V. Voltar")
                    print("0. Voltar ao menu principal")
                    escolha_lista = input(
                        "Escolha o número do paciente: ").strip().lower()
                    if escolha_lista == '0':
                        print("Voltando ao menu principal.\n")
                        return
                    if escolha_lista == 'v':
                        print()
                        break
                    if escolha_lista.isdigit():
                        escolha_int = int(escolha_lista)
                        if 1 <= escolha_int <= len(pacientes):
                            paciente = pacientes[escolha_int - 1]
                            print()
                            break
                    print("Seleção inválida. Tente novamente.")
                    continue
                print("Opção inválida. Tente novamente.")
            if paciente:
                break
            continue

        if len(candidatos) == 1:
            paciente = candidatos[0]
        else:
            print("Pacientes encontrados:")
            for idx, cand in enumerate(candidatos, start=1):
                idade_info = cand.get('idade')
                idade_str = f" | Idade: {idade_info}" if isinstance(
                    idade_info, (int, float)) else ""
                pid = obter_id_legivel(cand)
                print(f"{idx}. [{pid}] {cand.get('nome', '')}{idade_str}")
            print("V. Pesquisar novamente")
            print("0. Voltar ao menu principal")
            escolha = input("Escolha o número do paciente: ").strip().lower()
            if escolha == '0':
                print("Voltando ao menu principal.\n")
                return
            if escolha == 'v':
                continue
            if escolha.isdigit():
                escolha_int = int(escolha)
                if 1 <= escolha_int <= len(candidatos):
                    paciente = candidatos[escolha_int - 1]
                    break
            print("Seleção inválida. Tente novamente.")
            continue

    historico = [h for h in atendimentos if h.get('paciente', '').strip().lower()
                 == paciente['nome'].strip().lower()]
    if not historico:
        print("Nenhum atendimento registrado para este paciente.\n")
        return

    def _dt_value(entry):
        dh = entry.get('data_hora')
        if isinstance(dh, datetime):
            return dh
        if isinstance(dh, str):
            conv = validar_data_hora(dh)
            if conv:
                return conv
        return datetime.min

    historico_ordenado = sorted(historico, key=_dt_value)
    print(f"\n--- Histórico de {paciente['nome']} ---")
    for h in historico_ordenado:
        dh = h.get('data_hora')
        if isinstance(dh, datetime):
            dh_str = dh.strftime("%d/%m/%Y %H:%M")
        else:
            dh_str = str(dh)
        print(f"{dh_str} | Médico: {h.get('medico', '')} | {h.get('descricao', '')}")
    print()


def _buscar_medico_por_crm(crm_valor):
    alvo = (crm_valor or '').strip().lower()
    if not alvo:
        return None
    for medico in medicos:
        crm = str(medico.get('CRM') or '').strip().lower()
        if crm and crm == alvo:
            return medico
    return None


def autenticar_medico_por_crm(max_tentativas=3):
    if not medicos:
        print("Erro: não há médicos cadastrados.")
        return None

    tentativas = 0
    while tentativas < max_tentativas:
        crm_input = input("Informe seu CRM (ou 0 para voltar): ").strip()
        if crm_input == '0':
            print("Autenticação cancelada.\n")
            return None
        if not crm_input:
            print("CRM não pode ficar em branco.")
            tentativas += 1
            continue

        medico = _buscar_medico_por_crm(crm_input)
        if medico:
            print(
                f"Autenticação realizada. Bem-vindo(a), Dr(a). {medico.get('nome', '')}.\n")
            return medico

        tentativas += 1
        print("CRM não encontrado. Tente novamente.")

    print("Limite de tentativas atingido.\n")
    return None


def _obter_data_registro(registro):
    data_raw = registro.get('data_registro')
    if isinstance(data_raw, datetime):
        return data_raw
    if isinstance(data_raw, str):
        convertido = validar_data_hora(data_raw)
        if convertido:
            return convertido
    return None


def acessar_fila_atendimento_medico():
    medico = autenticar_medico_por_crm()
    if not medico:
        return

    if not cadastros_atendimento:
        print("Não há cadastros de atendimento registrados.\n")
        return

    registros_processados = []
    for registro in cadastros_atendimento:
        data_reg = _obter_data_registro(registro)
        registros_processados.append((data_reg, registro))

    if not registros_processados:
        print("Não há cadastros de atendimento registrados.\n")
        return

    registros_processados.sort(key=lambda item: item[0] or datetime.min)

    hoje = datetime.now().date()
    fila_hoje = []
    fila_antigos = []
    for data_reg, registro in registros_processados:
        if data_reg and data_reg.date() == hoje:
            fila_hoje.append((data_reg, registro))
        else:
            fila_antigos.append((data_reg, registro))

    def imprimir_registros(registros):
        for posicao, (data_reg, registro) in enumerate(registros, start=1):
            data_str = data_reg.strftime("%H:%M") if isinstance(
                data_reg, datetime) else "--:--"
            print(
                f"{posicao}. [{registro.get('paciente_id', 'Sem ID')}] {registro.get('paciente_nome', '')} | CPF: {registro.get('cpf', '')} | Convênio: {registro.get('convenio', '')} | Carteirinha: {registro.get('carteirinha', '')} | Registro: {data_str}")
        if registros:
            print()

    def processar_registro(registro, usar_paciente_da_fila):
        paciente = None
        if usar_paciente_da_fila:
            paciente = localizar_paciente_por_identificadores(
                registro.get('paciente_id'), registro.get('paciente_nome'))
            if paciente:
                print(
                    f"Iniciando registro de atendimento para {paciente.get('nome', '')}.\n")
            else:
                print(
                    "Paciente do cadastro não está nos registros. Iniciando atendimento manual.\n")
        kwargs = {
            'medico_autenticado': medico,
            'mostrar_agenda': False
        }
        if paciente:
            kwargs['paciente_preselecionado'] = paciente
        registrar_atendimento(**kwargs)
        if registro in cadastros_atendimento:
            cadastros_atendimento.remove(registro)

    houve_alteracao = False

    if fila_hoje:
        print(
            f"\nFila de cadastros aguardando atendimento em {hoje.strftime('%d/%m/%Y')}:")
        imprimir_registros(fila_hoje)
        for _, registro in fila_hoje:
            processar_registro(registro, usar_paciente_da_fila=True)
            houve_alteracao = True

    if fila_antigos:
        print("\nNenhum cadastro registrado hoje. Exibindo registros mais recentes:")
        imprimir_registros(fila_antigos)
        for _, registro in fila_antigos:
            processar_registro(registro, usar_paciente_da_fila=False)
            houve_alteracao = True

    if houve_alteracao:
        save_cadastros_atendimento()
        save_master_xlsx()


def ver_estatisticas():
    total = len(pacientes)
    print("\n--- Estatísticas de Pacientes ---")
    print(f"Total de pacientes: {total}")

    idades_validas = [p.get('idade') for p in pacientes
                      if isinstance(p.get('idade'), (int, float))]
    if idades_validas:
        media = sum(idades_validas) / len(idades_validas)
        media_inteira = round(media)
        pacientes_com_idade = [p for p in pacientes if isinstance(
            p.get('idade'), (int, float))]
        paciente_mais_novo = min(
            pacientes_com_idade, key=lambda registro: registro['idade'])
        paciente_mais_velho = max(
            pacientes_com_idade, key=lambda registro: registro['idade'])

        print(f"Idade média: {media_inteira} anos")
        print(
            f"Paciente mais novo: {paciente_mais_novo['nome']} ({paciente_mais_novo['idade']} anos)"
        )
        print(
            f"Paciente mais velho: {paciente_mais_velho['nome']} ({paciente_mais_velho['idade']} anos)"
        )
    else:
        print("Idade média: N/A (sem idades registradas)")
        print("Paciente mais novo: N/A")
        print("Paciente mais velho: N/A")
    print()


def simular_fila_pacientes():
    quantidade = 3
    fila = []
    for idx in range(1, quantidade + 1):
        nome = ""
        while not nome:
            nome = input(f"Nome do paciente #{idx}: ").strip()
            if not nome:
                print("O nome não pode ficar em branco.")

        cpf_formatado = None
        while cpf_formatado is None:
            cpf_input = input(
                f"CPF do paciente #{idx} (somente números): ").strip()
            cpf_digits = ''.join(ch for ch in cpf_input if ch.isdigit())
            if len(cpf_digits) == 11:
                cpf_formatado = formatar_cpf(cpf_digits)
            else:
                print("CPF inválido. Informe exatamente 11 dígitos.")

        fila.append({'nome': nome, 'cpf': cpf_formatado})

    if not fila:
        print("Nenhum paciente informado.\n")
        return

    atual = fila[0]
    print(
        f"\nAtendendo agora: {atual['nome']} (CPF {atual['cpf']})")
    if len(fila) > 1:
        print("Próximos pacientes na fila:")
        for posicao, paciente in enumerate(fila[1:], start=1):
            print(f"{posicao}. {paciente['nome']} (CPF {paciente['cpf']})")
    print()


def agendar(referencia_temporal=None, janela_minutos=None, paciente_preselecionado=None):
    if not pacientes:
        print("Erro: não há pacientes cadastrados.")
        return
    if not medicos:
        print("Erro: não há médicos cadastrados.")
        return
    if not disponibilidade:
        print("Erro: nenhum horário cadastrado. Utilize o módulo de disponibilidade antes de agendar.")
        return

    ref_dt = None
    janela_td = None
    janela_min_ativa = None
    if isinstance(referencia_temporal, datetime):
        ref_dt = referencia_temporal
    elif isinstance(referencia_temporal, str):
        ref_dt = validar_data_hora(referencia_temporal)
    if ref_dt:
        if isinstance(janela_minutos, (int, float)):
            janela_val = int(janela_minutos)
            if janela_val <= 0:
                janela_val = DEFAULT_JANELA_PROXIMIDADE_MINUTOS
            janela_min_ativa = janela_val
            janela_td = timedelta(minutes=janela_min_ativa)
            print(
                f"\nFiltrando horários dentro de {janela_min_ativa} minuto(s) a partir de {ref_dt.strftime('%d/%m/%Y %H:%M')}.")
        else:
            print(
                f"\nFiltrando horários a partir de {ref_dt.strftime('%d/%m/%Y %H:%M')}.")

    tipo = None
    exame_escolhido = ""
    while tipo is None:
        print("\nEscolha o tipo de atendimento:")
        print("1. Consulta")
        print("2. Exame")
        escolha_tipo = input(
            "Digite 1 para consulta ou 2 para exame: ").strip()
        if escolha_tipo == "1":
            tipo = "consulta"
        elif escolha_tipo == "2":
            tipo = "exame"
        else:
            print("Opção inválida. Tente novamente.")

    if tipo == "exame":
        exame_escolhido = selecionar_exame_interativo()
        if not exame_escolhido:
            return

    medicos_por_especialidade = {}
    for medico in medicos:
        nome_med = (medico.get('nome') or '').strip()
        if not nome_med:
            continue
        slots_med = slots_disponiveis_medico(nome_med, tipo)
        if ref_dt:
            slots_med = filtrar_slots_por_proximidade(
                slots_med, ref_dt, janela_td)
        if not slots_med:
            continue
        especialidade = (medico.get('especialidade')
                         or 'Não informado').strip() or 'Não informado'
        medicos_por_especialidade.setdefault(especialidade, []).append(medico)

    if not medicos_por_especialidade:
        print(
            "Erro: nenhum médico possui horários disponíveis para este tipo de atendimento.")
        return

    especialidades = sorted(
        medicos_por_especialidade.keys(), key=lambda esp: esp.lower())
    especialidade_selecionada = None
    medico_selecionado = None
    selected_slot = None

    while selected_slot is None:
        while especialidade_selecionada is None:
            print("\n--- Especialidades com horários disponíveis ---")
            for idx, esp in enumerate(especialidades, start=1):
                qtd = len(medicos_por_especialidade[esp])
                label = "médico" if qtd == 1 else "médicos"
                print(f"{idx}. {esp} ({qtd} {label})")
            print("0. Voltar ao menu principal")
            escolha_esp = input("Escolha a especialidade desejada: ").strip()
            if escolha_esp == '0':
                print("Voltando ao menu principal.\n")
                return
            if escolha_esp.isdigit():
                escolha_int = int(escolha_esp)
                if 1 <= escolha_int <= len(especialidades):
                    especialidade_selecionada = especialidades[escolha_int - 1]
                    break
            print("Seleção inválida. Tente novamente.")

        medicos_espec = sorted(
            medicos_por_especialidade[especialidade_selecionada],
            key=lambda registro: (registro.get('nome') or '').lower()
        )

        while medico_selecionado is None:
            print(f"\n--- Médicos de {especialidade_selecionada} ---")
            for idx, info in enumerate(medicos_espec, start=1):
                crm = info.get('CRM') or ''
                crm_str = f" | CRM: {crm}" if crm else ""
                print(f"{idx}. {info.get('nome', '')}{crm_str}")
            print("V. Escolher outra especialidade")
            print("0. Voltar ao menu principal")
            escolha_med = input("Escolha o médico desejado: ").strip().lower()
            if escolha_med == '0':
                print("Voltando ao menu principal.\n")
                return
            if escolha_med == 'v':
                especialidade_selecionada = None
                break
            if escolha_med.isdigit():
                escolha_int = int(escolha_med)
                if 1 <= escolha_int <= len(medicos_espec):
                    medico_selecionado = medicos_espec[escolha_int - 1]
                    break
            print("Seleção inválida. Tente novamente.")

        if medico_selecionado is None:
            continue

        nome_medico = (medico_selecionado.get('nome') or '').strip()
        slots_medico = slots_disponiveis_medico(nome_medico, tipo)
        if ref_dt:
            slots_medico = filtrar_slots_por_proximidade(
                slots_medico, ref_dt, janela_td)
        if not slots_medico:
            if janela_td:
                print(
                    "Nenhum horário dentro da janela selecionada para este médico. Escolha outro profissional.")
            else:
                print(
                    "Erro: médico selecionado não possui horários disponíveis no momento. Escolha outro profissional.")
            medico_selecionado = None
            continue

        datas_disponiveis = sorted(
            {slot['data_hora'].date() for slot in slots_medico})
        data_selecionada = None

        while data_selecionada is None:
            print(f"\n--- Datas disponíveis para {nome_medico} ---")
            for idx, data_opcao in enumerate(datas_disponiveis, start=1):
                qtd = sum(
                    1 for slot in slots_medico if slot['data_hora'].date() == data_opcao)
                label = "horário" if qtd == 1 else "horários"
                print(f"{idx}. {data_opcao.strftime('%d/%m/%Y')} ({qtd} {label})")
            print("V. Escolher outro médico")
            print("0. Voltar ao menu principal")
            escolha_data = input("Escolha a data desejada: ").strip().lower()
            if escolha_data == '0':
                print("Voltando ao menu principal.\n")
                return
            if escolha_data == 'v':
                medico_selecionado = None
                break
            if escolha_data.isdigit():
                escolha_int = int(escolha_data)
                if 1 <= escolha_int <= len(datas_disponiveis):
                    data_selecionada = datas_disponiveis[escolha_int - 1]
                    break
            print("Seleção inválida. Tente novamente.")

        if medico_selecionado is None:
            continue
        if data_selecionada is None:
            continue

        while True:
            horarios_no_dia = [
                slot for slot in slots_medico if slot['data_hora'].date() == data_selecionada]
            print(
                f"\n--- Horários em {data_selecionada.strftime('%d/%m/%Y')} ---")
            for idx, slot in enumerate(horarios_no_dia, start=1):
                print(f"{idx}. {slot['data_hora'].strftime('%H:%M')}")
            print("V. Escolher outra data")
            print("0. Voltar ao menu principal")
            print("I. Informar horário manualmente")
            escolha_hora = input(
                "Escolha o horário desejado: ").strip().lower()
            if escolha_hora == '0':
                print("Voltando ao menu principal.\n")
                return
            if escolha_hora == 'v':
                data_selecionada = None
                break
            if escolha_hora == 'i':
                hora_manual = input("Horário do atendimento (HH:MM): ").strip()
                try:
                    hora_dt = datetime.strptime(hora_manual, "%H:%M").time()
                except ValueError:
                    print("Formato de horário inválido. Tente novamente.")
                    continue
                data_hora_manual = datetime.combine(data_selecionada, hora_dt)
                slot_manual = next(
                    (slot for slot in slots_medico if slot['data_hora'] == data_hora_manual), None)
                if slot_manual:
                    selected_slot = slot_manual
                    break
                print("Erro: horário não cadastrado para este médico.")
                continue
            if escolha_hora.isdigit():
                escolha_int = int(escolha_hora)
                if 1 <= escolha_int <= len(horarios_no_dia):
                    selected_slot = horarios_no_dia[escolha_int - 1]
                    break
            print("Seleção inválida. Tente novamente.")

        if data_selecionada is None:
            continue
        if selected_slot:
            break

    paciente = paciente_preselecionado
    if not paciente:
        paciente = selecionar_paciente_interativo()
    if not paciente:
        return

    paciente_nome = paciente.get('nome', '').strip()
    medico_nome = (medico_selecionado.get('nome') or '').strip()
    data_hora = selected_slot['data_hora']

    conflito_apont = None
    for a in agendamentos:
        existing = a.get('data_hora')
        existing_dt = None
        if isinstance(existing, datetime):
            existing_dt = existing
        elif isinstance(existing, str):
            existing_dt = validar_data_hora(existing)

        if existing_dt and normalizar_texto(a.get('medico', '')) == normalizar_texto(medico_nome):
            if existing_dt == data_hora:
                conflito_apont = a
                break

    if conflito_apont:
        tipo_existente = (conflito_apont.get('tipo') or 'agendamento')
        article = 'uma' if tipo_existente == 'consulta' else 'um'
        print(
            f"Médico indisponível: já tem {article} {tipo_existente} neste horário.")
        return

    agendamentos.append({
        "paciente": paciente_nome,
        "medico": medico_nome,
        "data_hora": data_hora,
        "tipo": tipo,
        "exame": exame_escolhido
    })
    save_agendamentos()
    raw_slot = selected_slot.get('raw') if selected_slot else None
    slots_removidos = False
    if raw_slot and raw_slot in disponibilidade:
        disponibilidade.remove(raw_slot)
        slots_removidos = True

    medico_chave = normalizar_texto(medico_nome)
    for entry in list(disponibilidade):
        if normalizar_texto(entry.get('medico', '')) != medico_chave:
            continue
        dh_entry = entry.get('data_hora')
        if isinstance(dh_entry, str):
            dh_entry = validar_data_hora(dh_entry)
        if isinstance(dh_entry, datetime) and dh_entry == data_hora:
            disponibilidade.remove(entry)
            slots_removidos = True

    if slots_removidos:
        save_disponibilidade()
    save_master_xlsx()
    print("Agendamento realizado com sucesso!\n")


def listar_pacientes():
    print("\n--- Pacientes Cadastrados ---")
    if not pacientes:
        print("Nenhum paciente cadastrado.\n")
        return
    sincronizar_ids_pacientes()
    for p in pacientes:
        pid = obter_id_legivel(p)
        print(
            f"ID: {pid} | Nome: {p['nome']}, Idade: {p['idade']}, Telefone: {p['telefone']}")
    print()


def listar_medicos():
    print("\n--- Médicos Cadastrados ---")
    if not medicos:
        print("Nenhum médico cadastrado.\n")
        return
    for m in medicos:
        print(
            f"Nome: {m['nome']}, Especialidade: {m['especialidade']}, CRM: {m['CRM']}, Telefone: {m['telefone']}")
    print()


def listar_exames():
    print("\n--- Exames Cadastrados ---")
    if not exames:
        print("Nenhum exame cadastrado.\n")
        return
    for e in exames:
        print(
            f"Nome: {e['nome']}, Tipo: {e['tipo']}, Descrição: {e['descricao']}")
    print()


def listar_agendamentos():
    print("\n--- Agendamentos ---")
    if not agendamentos:
        print("Nenhum agendamento registrado.\n")
        return
    for a in agendamentos:
        dh = a.get('data_hora')
        if isinstance(dh, datetime):
            dh_str = dh.strftime("%d/%m/%Y %H:%M")
        else:
            dh_str = str(dh)
        print(
            f"Paciente: {a['paciente']}, Médico: {a['medico']}, Data/Hora: {dh_str}, Tipo: {a['tipo']}")
    print()


def consultar_agenda_medica_por_crm():
    if not medicos:
        print("Erro: não há médicos cadastrados.")
        return

    crm_input = input("Informe o CRM do médico: ").strip()
    if not crm_input:
        print("CRM não pode ficar em branco.\n")
        return

    medico = _buscar_medico_por_crm(crm_input)
    if not medico:
        print("CRM não encontrado.\n")
        return

    data_filtro = None
    data_input = input(
        "Data para consulta (DD/MM/AAAA) - deixe em branco para todas: ").strip()
    if data_input:
        try:
            data_filtro = datetime.strptime(data_input, "%d/%m/%Y").date()
        except ValueError:
            print("Data inválida. Use o formato DD/MM/AAAA.\n")
            return

    nome_medico = medico.get('nome', '')
    alvo_norm = normalizar_texto(nome_medico)
    agenda_filtrada = []
    for ag in agendamentos:
        if normalizar_texto(ag.get('medico', '')) != alvo_norm:
            continue
        dh = ag.get('data_hora')
        if isinstance(dh, str):
            dh = validar_data_hora(dh)
        if not isinstance(dh, datetime):
            continue
        if data_filtro and dh.date() != data_filtro:
            continue
        agenda_filtrada.append((dh, ag))

    if not agenda_filtrada:
        if data_filtro:
            print("Nenhum agendamento encontrado para esta data.\n")
        else:
            print("Nenhum agendamento encontrado para este médico.\n")
        return

    agenda_filtrada.sort(key=lambda item: item[0])
    print(f"\nAgenda do(a) Dr(a). {nome_medico}:")
    if data_filtro:
        print(f"Data filtrada: {data_filtro.strftime('%d/%m/%Y')}")
    for dh, ag in agenda_filtrada:
        tipo = ag.get('tipo') or 'indefinido'
        exame = ag.get('exame') or ''
        exame_str = f" | Exame: {exame}" if exame else ''
        print(
            f"{dh.strftime('%d/%m/%Y %H:%M')} | Paciente: {ag.get('paciente', '')} | Tipo: {tipo}{exame_str}")
    print()


def listar_disponibilidade_por_medico():
    print("\n--- Agenda de Horários ---")
    if not disponibilidade:
        print("Nenhum horário cadastrado.\n")
        return

    medicos_disp = sorted({(slot.get('medico') or '').strip()
                          for slot in disponibilidade if slot.get('medico')})
    medicos_disp = [nome for nome in medicos_disp if nome]
    if not medicos_disp:
        print("Nenhum médico com horários cadastrados.\n")
        return

    while True:
        termo = input("Nome (ou parte do nome) do médico: ").strip()
        if not termo:
            print("Informe ao menos uma letra para buscar o médico.")
            continue

        termo_norm = normalizar_texto(termo)
        candidatos = [nome for nome in medicos_disp if termo_norm
                      in normalizar_texto(nome)]

        if not candidatos:
            print("Nenhum médico encontrado com esse termo.")
            while True:
                print("1. Listar médicos com horários cadastrados")
                print("0. Voltar ao menu anterior")
                acao = input("Escolha uma opção: ").strip()
                if acao == "0":
                    print()
                    return
                if acao == "1":
                    print("\n--- Médicos com horários cadastrados ---")
                    for idx, nome in enumerate(medicos_disp, start=1):
                        print(f"{idx}. {nome}")
                    print("0. Voltar")
                    escolha = input(
                        "Escolha o número do médico (ou 0 para voltar): ").strip()
                    if escolha.isdigit():
                        escolha_int = int(escolha)
                        if escolha_int == 0:
                            print()
                            break
                        if 1 <= escolha_int <= len(medicos_disp):
                            mostrar_agenda_medico(
                                medicos_disp[escolha_int - 1])
                            return
                    print("Seleção inválida. Tente novamente.")
                    continue
                print("Opção inválida. Tente novamente.")
            continue

        while True:
            print("Médicos encontrados:")
            for idx, nome in enumerate(candidatos, start=1):
                print(f"{idx}. {nome}")
            print("V. Pesquisar novamente")
            print("0. Voltar ao menu anterior")
            escolha_cand = input(
                "Escolha o número do médico: ").strip().lower()
            if escolha_cand == '0':
                print()
                return
            if escolha_cand == 'v':
                break
            if escolha_cand.isdigit():
                escolha_int = int(escolha_cand)
                if 1 <= escolha_int <= len(candidatos):
                    mostrar_agenda_medico(candidatos[escolha_int - 1])
                    return
            print("Seleção inválida. Tente novamente.")
        continue


def mostrar_agenda_medico(medico_nome):
    alvo_norm = normalizar_texto(medico_nome)
    slots = []
    for slot in disponibilidade:
        nome_slot = (slot.get('medico') or '').strip()
        if not nome_slot:
            continue
        if normalizar_texto(nome_slot) != alvo_norm:
            continue
        dh = slot.get('data_hora')
        if isinstance(dh, str):
            conv = validar_data_hora(dh)
            dh = conv
        slots.append({
            'data_hora': dh,
            'tipo': (slot.get('tipo') or '').lower() or 'indefinido'
        })

    if not slots:
        print(f"Médico {medico_nome} não possui horários cadastrados.\n")
        return

    slots_ordenados = sorted(
        slots, key=lambda entry: entry['data_hora'] or datetime.max)
    print(f"\n--- Horários de {medico_nome} ---")
    for s in slots_ordenados:
        dh = s['data_hora']
        if isinstance(dh, datetime):
            dh_str = dh.strftime("%d/%m/%Y %H:%M")
        elif dh:
            dh_str = str(dh)
        else:
            dh_str = "Data/Hora indefinida"
        tipo_str = s['tipo'] or 'indefinido'
        print(f"{dh_str} | Tipo: {tipo_str}")
    print()


def cancelar_agenda_medico_por_data():
    if not disponibilidade:
        print("Nenhum horário cadastrado para cancelar.\n")
        return

    medicos_disp = sorted({(slot.get('medico') or '').strip()
                          for slot in disponibilidade if slot.get('medico')})
    medicos_disp = [nome for nome in medicos_disp if nome]
    if not medicos_disp:
        print("Nenhum médico com horários cadastrados.\n")
        return

    medico_escolhido = None
    while medico_escolhido is None:
        termo = input("Nome (ou parte do nome) do médico: ").strip()
        if not termo:
            print("Informe ao menos uma letra para buscar o médico.")
            continue
        termo_norm = normalizar_texto(termo)
        candidatos = [nome for nome in medicos_disp if termo_norm
                      in normalizar_texto(nome)]
        if not candidatos:
            print("Nenhum médico encontrado com esse termo.")
            while True:
                print("1. Listar médicos com horários cadastrados")
                print("0. Voltar ao menu anterior")
                acao = input("Escolha uma opção: ").strip()
                if acao == "0":
                    print()
                    return
                if acao == "1":
                    print("\n--- Médicos com horários cadastrados ---")
                    for idx, nome in enumerate(medicos_disp, start=1):
                        print(f"{idx}. {nome}")
                    print("0. Voltar")
                    escolha = input("Escolha o número do médico: ").strip()
                    if escolha.isdigit():
                        escolha_int = int(escolha)
                        if escolha_int == 0:
                            print()
                            break
                        if 1 <= escolha_int <= len(medicos_disp):
                            medico_escolhido = medicos_disp[escolha_int - 1]
                            print()
                            break
                    print("Seleção inválida. Tente novamente.")
                    continue
                print("Opção inválida. Tente novamente.")
            continue

        if len(candidatos) == 1:
            medico_escolhido = candidatos[0]
        else:
            print("Médicos encontrados:")
            for idx, nome in enumerate(candidatos, start=1):
                print(f"{idx}. {nome}")
            print("0. Pesquisar novamente")
            print("M. Voltar ao menu anterior")
            escolha = input("Escolha o número do médico: ").strip()
            if escolha.lower() == 'm':
                print()
                return
            if escolha.isdigit():
                escolha_int = int(escolha)
                if escolha_int == 0:
                    continue
                if 1 <= escolha_int <= len(candidatos):
                    medico_escolhido = candidatos[escolha_int - 1]
                    break
            print("Seleção inválida. Tente novamente.")

    data_base = None
    while data_base is None:
        data_str = input("Data a cancelar (DD/MM/AAAA): ").strip()
        if not data_str:
            print("Informe uma data válida.")
            continue
        try:
            data_base = datetime.strptime(data_str, "%d/%m/%Y")
        except ValueError:
            print("Formato de data inválido. Tente novamente.")
            continue

    alvo_norm = normalizar_texto(medico_escolhido)
    slots_remover = []
    for slot in disponibilidade:
        nome_slot = (slot.get('medico') or '').strip()
        if not nome_slot or normalizar_texto(nome_slot) != alvo_norm:
            continue
        dh = slot.get('data_hora')
        if isinstance(dh, str):
            dh = validar_data_hora(dh)
        if isinstance(dh, datetime) and dh.date() == data_base.date():
            slots_remover.append(slot)

    if not slots_remover:
        print(
            f"Nenhum horário encontrado para {medico_escolhido} em {data_base.strftime('%d/%m/%Y')}.")
        return

    print(
        f"Encontrados {len(slots_remover)} horário(s) para {medico_escolhido} em {data_base.strftime('%d/%m/%Y')}.")
    confirmar = input("Confirmar cancelamento? (S/N): ").strip().lower()
    if confirmar != 's':
        print("Cancelamento abortado.\n")
        return

    for slot in slots_remover:
        if slot in disponibilidade:
            disponibilidade.remove(slot)

    save_disponibilidade()
    save_master_xlsx()
    print("Horários cancelados com sucesso.\n")


def menu_cadastro():
    while True:
        print("\n--- Módulo de Cadastro ---")
        print("1. Cadastrar paciente")
        print("2. Cadastrar médico")
        print("3. Cadastrar exame")
        print("4. Listar pacientes")
        print("5. Listar médicos")
        print("6. Listar exames")
        print("7. Cadastrar convênio")
        print("8. Listar convênios")
        print("0. Voltar ao menu principal")
        opcao = input("Escolha uma opção: ").strip()

        if opcao == "1":
            cadastrar_paciente()
        elif opcao == "2":
            cadastrar_medico()
        elif opcao == "3":
            cadastrar_exame()
        elif opcao == "4":
            listar_pacientes()
        elif opcao == "5":
            listar_medicos()
        elif opcao == "6":
            listar_exames()
        elif opcao == "7":
            cadastrar_convenio()
        elif opcao == "8":
            listar_convenios()
        elif opcao == "0":
            print("Voltando ao menu principal.\n")
            break
        else:
            print("Opção inválida. Tente novamente.\n")


def menu_agendamento():
    while True:
        print("\n--- Módulo de Agendamento ---")
        print("1. Realizar agendamento")
        print("2. Listar agendamentos")
        print("0. Voltar ao menu principal")
        opcao = input("Escolha uma opção: ").strip()

        if opcao == "1":
            agendar()
        elif opcao == "2":
            listar_agendamentos()
        elif opcao == "0":
            print("Voltando ao menu principal.\n")
            break
        else:
            print("Opção inválida. Tente novamente.\n")


def menu_agenda_horarios():
    while True:
        print("\n--- Agenda Horários de Exames/Médicos ---")
        print("1. Cadastrar/Atualizar disponibilidade")
        print("2. Consultar horários por médico")
        print("3. Cancelar horários por médico/data")
        print("0. Voltar ao menu principal")
        opcao = input("Escolha uma opção: ").strip()

        if opcao == "1":
            cadastrar_disponibilidade()
        elif opcao == "2":
            listar_disponibilidade_por_medico()
        elif opcao == "3":
            cancelar_agenda_medico_por_data()
        elif opcao == "0":
            print("Voltando ao menu principal.\n")
            break
        else:
            print("Opção inválida. Tente novamente.\n")


def menu_cadastro_atendimento():
    while True:
        print("\n--- Cadastro de Atendimento ---")
        print("1. Registrar cadastro de atendimento")
        print("2. Buscar cadastros por paciente")
        print("0. Voltar ao menu principal")
        opcao = input("Escolha uma opção: ").strip()

        if opcao == "1":
            cadastrar_atendimento_paciente()
        elif opcao == "2":
            buscar_cadastros_atendimento_por_nome()
        elif opcao == "0":
            print("Voltando ao menu principal.\n")
            break
        else:
            print("Opção inválida. Tente novamente.\n")


def menu_modulo_medico():
    while True:
        print("\n--- Módulo Médico ---")
        print("1. Atendimento")
        print("2. Ver histórico de paciente")
        print("0. Voltar ao menu principal")
        opcao = input("Escolha uma opção: ").strip()

        if opcao == "1":
            acessar_fila_atendimento_medico()
        elif opcao == "2":
            visualizar_historico_paciente()
        elif opcao == "0":
            print("Voltando ao menu principal.\n")
            break
        else:
            print("Opção inválida. Tente novamente.\n")


def menu():
    while True:
        print("=== CLÍNICA VIDA+ ===")
        print("1. Cadastro")
        print("2. Agendamento")
        print("3. Cadastro Atendimento")
        print("4. Agenda Horários de Exames/Médicos")
        print("5. Módulo Médico")
        print("6. Ver estatísticas")
        print("7. Sair")
        opcao = input("Escolha uma opção: ").strip()

        if opcao == "1":
            menu_cadastro()
        elif opcao == "2":
            menu_agendamento()
        elif opcao == "3":
            menu_cadastro_atendimento()
        elif opcao == "4":
            menu_agenda_horarios()
        elif opcao == "5":
            menu_modulo_medico()
        elif opcao == "6":
            ver_estatisticas()
        elif opcao == "7":
            print("Encerrando o sistema...")
            break
        else:
            print("Opção inválida. Tente novamente.\n")


if __name__ == "__main__":
    menu()
